"""
Tests for the three format writers and the registry that resolves them.

Every case runs against all three formats through the same parametrised test, because the
contract is the point: ``base/`` knows nothing about any format beyond ``write_part`` and
``merge_parts``, so anything true of one must be true of all three or the graph would need
to know which is which.

The claim being tested is narrow and total: **an export of N records contains N records,
once each, in order, with one header.** Every format gets it a different way — CSV
concatenates bytes and drops repeated headers, Parquet appends row groups, XLSX rewrites a
workbook — and a merge that produced a valid file with 50 records missing would pass any
looser check.

The all-NULL column has its own test because it is the one that breaks Parquet: a batch
whose values are all NULL infers a null column, which cannot then hold batch 2's strings,
and the export would fail thousands of records in on a query that works everywhere else.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Callable, Dict, List

import pytest
from litestar.exceptions import HTTPException

from app.models.downloader_agents import FORMAT_CSV, FORMAT_PARQUET, FORMAT_XLS
from app.services.downloader_agents.base.part_writer import (
    extension_for,
    forget_export_caches,
    writer_for,
)

FORMATS = [FORMAT_CSV, FORMAT_XLS, FORMAT_PARQUET]


def _records(start: int, count: int, note: Any = None) -> List[Dict[str, Any]]:
    """``count`` records numbered from ``start``."""
    return [
        {"id": index, "name": f"n{index}", "qty": index % 7, "note": note}
        for index in range(start, start + count)
    ]


def _read_back(path: Path, file_format: str) -> List[Dict[str, Any]]:
    """
    The records in a finished artifact, whatever format it is.

    Reads each format with the library a *user* would open it with, not with the writer
    that produced it — a round trip through our own code could agree with itself about a
    file nothing else can read.
    """
    if file_format == FORMAT_CSV:
        with open(path, newline="", encoding="utf-8") as handle:
            return list(csv.DictReader(handle))

    if file_format == FORMAT_XLS:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            return []

        header = [str(name) for name in rows[0]]
        return [dict(zip(header, row)) for row in rows[1:]]

    import pyarrow.parquet as pq

    return pq.read_table(path).to_pylist()


@pytest.fixture
def export_dirs(tmp_path: Path) -> Callable:  # noqa: ANN001
    """A ``parts/`` directory and an artifact path, per format."""

    def _make(file_format: str) -> tuple:
        root = tmp_path / file_format
        parts = root / "parts"
        parts.mkdir(parents=True, exist_ok=True)
        return parts, root / f"artifact{extension_for(file_format)}"

    return _make


@pytest.fixture(autouse=True)
def _clear_format_caches(tmp_path: Path):  # noqa: ANN001, ANN202
    """
    Drop any pinned Parquet schema between tests.

    The cache is keyed by parts directory and ``tmp_path`` differs per test, so this is
    belt and braces — but a leaked schema would make one test's types decide another's,
    which is the sort of failure that only shows up when the suite is reordered.
    """
    yield
    for path in tmp_path.rglob("parts"):
        forget_export_caches(path)


# ---- The round trip ----

@pytest.mark.parametrize("file_format", FORMATS)
class TestRoundTrip:
    async def test_three_batches_merge_into_one_complete_file(
        self, file_format: str, export_dirs: Callable,
    ) -> None:
        parts_dir, artifact = export_dirs(file_format)
        writer = writer_for(file_format)
        extension = extension_for(file_format)

        batches = [_records(1, 50), _records(51, 50), _records(101, 25)]
        paths = []
        written = 0

        for number, batch in enumerate(batches, start=1):
            path = parts_dir / f"part-{number:06d}{extension}"
            written += await writer.write_part(batch, path)
            paths.append(path)

        merged = await writer.merge_parts(paths, artifact)

        assert written == 125
        assert merged == 125

        records = _read_back(artifact, file_format)
        assert len(records) == 125
        # Order and identity, not just the count: parts merged out of order produce a
        # valid file with the right rows in the wrong sequence.
        assert [int(row["id"]) for row in records] == list(range(1, 126))
        # One header, not three.
        assert [str(row["name"]) for row in records][:2] == ["n1", "n2"]

    async def test_an_empty_batch_writes_no_file(
        self, file_format: str, export_dirs: Callable,
    ) -> None:
        """
        "A part file exists" has to mean "a batch had rows in it".

        The merge and the cleanup both assume it, and an empty part would be merged as a
        header with no records — or counted as a part that was never written.
        """
        parts_dir, _artifact = export_dirs(file_format)
        writer = writer_for(file_format)
        path = parts_dir / f"part-000001{extension_for(file_format)}"

        assert await writer.write_part([], path) == 0
        assert not path.exists()

    async def test_no_parts_still_produces_a_readable_file(
        self, file_format: str, export_dirs: Callable,
    ) -> None:
        """
        A query that matched nothing is an answer, not a failure.

        The file has to exist: an empty result with no artifact is a download link that
        404s, which reads as a broken application rather than as "no records".
        """
        _parts_dir, artifact = export_dirs(file_format)
        writer = writer_for(file_format)

        assert await writer.merge_parts([], artifact) == 0
        assert artifact.is_file()
        assert _read_back(artifact, file_format) == []

    async def test_a_single_batch_merges_unchanged(
        self, file_format: str, export_dirs: Callable,
    ) -> None:
        parts_dir, artifact = export_dirs(file_format)
        writer = writer_for(file_format)
        path = parts_dir / f"part-000001{extension_for(file_format)}"

        await writer.write_part(_records(1, 7), path)
        merged = await writer.merge_parts([path], artifact)

        assert merged == 7
        assert [int(row["id"]) for row in _read_back(artifact, file_format)] == list(
            range(1, 8)
        )

    async def test_an_all_null_column_survives_a_later_value(
        self, file_format: str, export_dirs: Callable,
    ) -> None:
        """
        Batch 1's column is entirely NULL; batch 2's has text in it.

        This is the case that breaks a typed format: inferring a null column from batch 1
        leaves a schema that cannot hold batch 2, and the export dies partway through a
        query that is perfectly valid.
        """
        parts_dir, artifact = export_dirs(file_format)
        writer = writer_for(file_format)
        extension = extension_for(file_format)

        first = parts_dir / f"part-000001{extension}"
        second = parts_dir / f"part-000002{extension}"

        await writer.write_part(_records(1, 3, note=None), first)
        await writer.write_part(_records(4, 3, note="seen"), second)

        merged = await writer.merge_parts([first, second], artifact)

        assert merged == 6

        records = _read_back(artifact, file_format)
        assert [int(row["id"]) for row in records] == [1, 2, 3, 4, 5, 6]
        assert [row["note"] for row in records][3:] == ["seen", "seen", "seen"]


# ---- The registry ----

class TestWriterRegistry:
    @pytest.mark.parametrize(
        ("file_format", "extension"),
        [(FORMAT_CSV, ".csv"), (FORMAT_XLS, ".xlsx"), (FORMAT_PARQUET, ".parquet")],
    )
    def test_the_extension_matches_the_writer(
        self, file_format: str, extension: str,
    ) -> None:
        """
        ``extension_for`` is read without importing the writer, so the two could drift —
        and a part file named ``.csv`` written by the Parquet writer would fail its merge
        for a reason having nothing to do with the data.
        """
        assert extension_for(file_format) == extension
        assert writer_for(file_format).extension == extension

    @pytest.mark.parametrize("file_format", FORMATS)
    def test_every_writer_declares_a_media_type(self, file_format: str) -> None:
        """The download route sends it as Content-Type, so it cannot be empty."""
        assert writer_for(file_format).media_type

    def test_an_unknown_format_is_refused_by_name(self) -> None:
        with pytest.raises(HTTPException) as excinfo:
            writer_for("pdf")

        assert excinfo.value.status_code == 400
        assert "pdf" in str(excinfo.value.detail)
        assert "csv" in str(excinfo.value.detail)

    def test_a_writer_is_resolved_once(self) -> None:
        """Importing pyarrow per batch would be a real cost on a long export."""
        assert writer_for(FORMAT_PARQUET) is writer_for(FORMAT_PARQUET)
