"""
Excel parts, and merging them into one workbook.

Implements the contract in
:mod:`app.services.downloader_agents.base.part_writer`. Nothing else imports this
module — the registry there resolves it by format name.

**The folder is ``xls`` and the files are ``.xlsx``.** The folder name is the format as
people ask for it; the extension is what actually gets written. Legacy ``.xls`` (BIFF)
is not produced at all: it caps out at 65,536 rows, which an export whose whole purpose
is "more records than fit in a chat message" would hit routinely, and openpyxl cannot
write it. Anything asking for "xls" gets a real modern workbook.

**Why openpyxl in write-only mode.** An ordinary ``Workbook()`` holds every cell as a
Python object until it is saved — the whole export in memory, which is the thing this
feature is built to avoid. ``write_only=True`` streams rows to the sheet as they are
appended and keeps almost nothing, at the cost of the workbook being append-only. That
cost is free here: an export is written once, front to back.

**Why the merge reads the parts back.** This is the one format where merging is not a
copy. An ``.xlsx`` file is a zip archive of XML parts with a shared string table and a
declared dimension — two of them cannot be concatenated into a valid third. So the
merge opens each part read-only and streams its rows into one new workbook. That is
more expensive than the CSV path and it is unavoidable; what makes it acceptable is
``read_only=True``, which streams the source rows rather than loading each part whole.

**Why parts are ``.xlsx`` too, and not CSV.** A part file in the target format keeps
one property that matters when an export fails: what is left on disk is readable with
the tool the user was going to open the result with. It also keeps each format's part
and merge logic in one module, which is why the packages are split by format at all.
"""

import asyncio
from pathlib import Path
from typing import Any, Dict, List

# Module scope, not inside the functions below — for the same reason spelled out in
# app/services/downloader_agents/parquet/parquet_writer.py: a C extension first imported
# on a worker thread that is later destroyed can take the process down on its next use.
# openpyxl is far less prone to it than pyarrow, and being consistent about it costs
# nothing: `base/part_writer` only imports this module when a caller asks for Excel.
from openpyxl import Workbook, load_workbook

#: Part of the PartWriter contract.
extension = ".xlsx"
media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# The sheet every export writes into. Named rather than left as openpyxl's default
# "Sheet" because it is what the user sees on the tab.
_SHEET_TITLE = "Records"

# Cell types openpyxl can write directly. Anything else — a date, a Decimal, a UUID, a
# dict from a JSON column — is written as its string form, which is what the CSV and
# the agent's own answer already show. Excel would otherwise refuse the whole workbook
# over one unexpected value.
_NATIVE_TYPES = (str, int, float, bool)


async def write_part(rows: List[Dict[str, Any]], path: Path) -> int:
    """
    Write one batch as a single-sheet workbook with a header row.

    Same shape as the CSV writer: header from the first row's keys, every part
    independently readable, nothing written at all for an empty batch.
    """
    if not rows:
        return 0

    def _write() -> int:
        fieldnames = list(rows[0].keys())

        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title=_SHEET_TITLE)

        sheet.append(fieldnames)

        for row in rows:
            sheet.append(_row_cells(row, fieldnames))

        try:
            workbook.save(path)
        finally:
            # A write-only workbook holds an open archive handle. Not closing it leaks
            # a file descriptor per batch, which over ten thousand batches is an
            # export that dies of something entirely unrelated to the data.
            workbook.close()

        return len(rows)

    return await asyncio.to_thread(_write)


async def merge_parts(paths: List[Path], destination: Path) -> int:
    """
    Stream every part's rows into one workbook.

    The first part supplies the header; later parts have theirs skipped, exactly as in
    the CSV merge. Returns the number of data rows written, counted as they go past.

    Each source is opened with ``read_only=True`` and closed straight after, so at no
    point is more than one part's worth of rows resident.
    """

    def _merge() -> int:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title=_SHEET_TITLE)
        rows_written = 0
        width = 0

        try:
            for index, part in enumerate(paths):
                copied, width = _copy_part(
                    part, sheet, keep_header=index == 0, width=width,
                )
                rows_written += copied

            if not width:
                # No parts at all: the query matched nothing. An empty sheet is still
                # written, because a missing file is a broken download and "no
                # records" is a truthful result.
                sheet.append([])

            workbook.save(destination)
        finally:
            workbook.close()

        return rows_written

    return await asyncio.to_thread(_merge)


def _copy_part(part: Path, sheet: Any, keep_header: bool, width: int) -> tuple:
    """
    Stream one part's rows into the open sheet. Returns ``(rows_copied, width)``.

    ``width`` is the header's column count, carried between parts so every row can be
    padded to it. Padding matters for the same reason :func:`_row_cells` exists — a row
    saved narrower than the header reads back with its last field missing — and here it
    also guards against a part written by an older version of this module.
    """
    source = load_workbook(part, read_only=True, data_only=True)
    rows_copied = 0

    try:
        for index, row in enumerate(source[source.sheetnames[0]].iter_rows(
            values_only=True,
        )):
            values = list(row)

            if index == 0:
                if keep_header:
                    width = len(values)
                    sheet.append(values)
                continue

            sheet.append(_rectangular(values, width))
            rows_copied += 1
    finally:
        # read_only workbooks keep the zip open until closed, and the merge opens one
        # per part.
        source.close()

    return rows_copied, width


def _row_cells(row: Dict[str, Any], fieldnames: List[str]) -> List[Any]:
    """One record as a full-width row of cells. See :func:`_rectangular`."""
    return _rectangular(
        [_cell(row.get(name)) for name in fieldnames], len(fieldnames),
    )


def _rectangular(values: List[Any], width: int) -> List[Any]:
    """
    One row, padded to ``width``, with a cell that actually exists in the last column.

    **openpyxl's write-only mode creates no cell at all for a ``None``.** So a record whose
    final column is NULL is *saved narrower than the header*, and read back as a shorter
    tuple — which makes that column disappear from the row entirely. Anything zipping the
    header to the row (this module's own merge, pandas, a person's script) then drops the
    field, or with two such columns misattributes the values.

    Both the part writer and the merge go through here, and the merge needs it as much as
    the writer does: it reads a full-width row off a part and would otherwise re-drop the
    trailing NULL on the way out, undoing the fix one function earlier.

    A trailing NULL becomes an empty cell. That costs the ability to tell NULL from an empty
    string in the final column — the same trade the CSV export already makes for every
    column — and it buys a rectangular sheet. Interior NULLs stay ``None``, a genuinely
    empty cell with its column's type intact, because only trailing cells are trimmed.
    """
    padded = list(values)

    if width and len(padded) < width:
        padded.extend([""] * (width - len(padded)))

    if padded and padded[-1] is None:
        padded[-1] = ""

    return padded


def _cell(value: Any) -> Any:
    """
    One value as something openpyxl will accept.

    ``None`` stays ``None`` — an empty cell, which is what a NULL is — and the four
    native types pass through so numbers stay numbers and sort as numbers in Excel.
    Everything else becomes ``str(value)``, matching what the CSV writer and the
    agent's own answer show for the same row.
    """
    if value is None or isinstance(value, _NATIVE_TYPES):
        return value

    return str(value)
